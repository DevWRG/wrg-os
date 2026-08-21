#!/usr/bin/env python3
"""Ubah workbook "Hasil Perhitungan KSO Per Tes" jadi JSON siap-impor (migrasi 097).

KENAPA DUA LANGKAH (python -> JSON -> node -> DB), BUKAN LANGSUNG:
1. Repo ini PUBLIK. Nama faskes, nomor MOU, dan SN alat tidak boleh masuk git. JSON hasil
   konversi WAJIB ditulis ke luar repo (skrip menolak menulis ke dalam working tree).
2. apps/api tidak punya dependensi pembaca xlsx, dan menambahkannya cuma untuk skrip ops
   sekali jalan tidak sepadan (belum lagi aturan pnpm minimumReleaseAge). openpyxl sudah
   ada di jalur python yang sama dengan services/ai.

CARA PAKAI:
    # 1. Export spreadsheet-nya ke .xlsx (File > Download > Microsoft Excel)
    # 2. Konversi:
    python3 scripts/ops/kso-sheet-to-json.py ~/Downloads/kso.xlsx --out ~/kso-import.json
    # 3. Impor (lihat scripts/ops/kso-asset-import.mjs):
    node scripts/ops/kso-asset-import.mjs --file ~/kso-import.json          # pratinjau
    node scripts/ops/kso-asset-import.mjs --file ~/kso-import.json --apply  # tulis

Skrip ini murni transformasi + audit: tidak menyentuh database sama sekali.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl belum terpasang. Jalankan: python3 -m pip install openpyxl")

SHEET_POPULASI = "Populasi KSO"
SHEET_TES_2026 = "2026 KSO Tes"
SHEET_REAGENT_2026 = "2026 KSO Reagent"
SHEET_TES_2025 = "2025 KSO Tes"
SHEET_REAGENT_2025 = "2025 KSO Reagent"
SHEET_PARAM_2026 = "rekap perparameterkimia 2026"

# Peta kolom bulanan per sheet. Ditulis eksplisit, BUKAN ditebak dari nama kolom, karena
# penamaannya kacau dan tidak konsisten antar sheet: "January" vs "Jumlah Tes Januari"
# vs "TES MARET", dan sheet 2025 punya ekor kolom 2026 yang harus diabaikan (sheet 2026
# yang berwenang untuk tahun 2026 — kalau dua-duanya dipakai, yang belakangan menimpa
# yang benar tanpa jejak).
MONTH_COLUMNS = {
    SHEET_TES_2026: {
        "January": (2026, 1), "February": (2026, 2), "Maret": (2026, 3),
        "April": (2026, 4), "Mei": (2026, 5), "JUNI": (2026, 6),
        "Juli": (2026, 7), "Agustus": (2026, 8), "September": (2026, 9),
        "Oktober": (2026, 10), "November": (2026, 11), "Desember": (2026, 12),
    },
    SHEET_REAGENT_2026: {
        "Jumlah Tes Januari 2026": (2026, 1), "Jumlah Tes Februari 2026": (2026, 2),
        "TES MARET": (2026, 3), "TES APRIL": (2026, 4), "TES MEI": (2026, 5),
    },
    SHEET_TES_2025: {
        f"Jumlah Tes {n}": (2025, i)
        for i, n in enumerate(
            ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
             "Juli", "Agustus", "September", "Oktober", "November", "Desember"], 1)
    },
    SHEET_REAGENT_2025: {
        f"Jumlah Tes {n}": (2025, i)
        for i, n in enumerate(
            ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
             "Juli", "Agustus", "September", "Oktober", "November", "Desember"], 1)
    },
}

BULAN_KE_NOMOR = {
    "JANUARI": 1, "JANUARY": 1, "FEBRUARI": 2, "FEBRUARY": 2, "MARET": 3, "MARCH": 3,
    "APRIL": 4, "MEI": 5, "MAY": 5, "JUNI": 6, "JUNE": 6, "JULI": 7, "JULY": 7,
    "AGUSTUS": 8, "AUGUST": 8, "SEPTEMBER": 9, "OKTOBER": 10, "OCTOBER": 10,
    "NOVEMBER": 11, "DESEMBER": 12, "DECEMBER": 12,
}

# Kolom parameter di sheet rekap per-parameter, apa adanya -> nama kanonik.
PARAM_COLUMNS = {
    "ALB": "Albumin", "ALP": "ALP", "AMILASE": "Amilase", "BIL D": "Bilirubin D",
    "BIL T": "Bilirubin T", "COLESTEROL": "Cholesterol", "UREUM": "Ureum",
    "CREAT": "Creatinin", "GLUCOSE": "Glucose", "HDL": "HDL", "LDL": "LDL",
    "SGOT": "SGOT", "SGPT": "SGPT", "TG": "Trigliserida", "COLINES": "Cholinesterase",
    "GGT": "GGT", "TP": "Total Protein", "LDH": "LDH", "UA": "Uric Acid",
    "SI": "Serum Iron", "TIBC": "TIBC", "GLOBOLIN": "Globulin",
}


def norm_sn(raw) -> str:
    """SN kanonik. Lihat komentar `sn_key` di migrasi 097 — normalisasi ini yang
    menyatukan `00829` / `829` / `5360.0` jadi satu alat."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    # openpyxl mengembalikan SN numerik sebagai float -> "5360.0"
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = s.upper().replace(" ", "")
    # lstrip('0') pada SN yang isinya nol semua akan jadi string kosong -> pertahankan asli
    return s.lstrip("0") or s


def slug(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def as_number(v):
    """Angka bulanan, atau None kalau selnya bukan angka.

    Sel non-angka di sheet ini bukan cuma kosong: ada '#N/A' (VLOOKUP gagal), 'UNITY',
    'DITARIK' (alat ditarik), dan 'beli reagen' (skema berubah di tengah tahun). Semua
    itu jadi NULL, tapi dihitung di laporan supaya tidak lenyap tanpa jejak."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def as_int(v):
    n = as_number(v)
    return int(n) if n is not None else None


def as_date(v):
    """Masa berlaku MOU ditulis dua gaya: datetime betulan dan teks '03 Mei 2024'."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$", s)
    if m:
        bulan = BULAN_KE_NOMOR.get(m.group(2).upper())
        if bulan:
            try:
                return date(int(m.group(3)), bulan, int(m.group(1))).isoformat()
            except ValueError:
                return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):  # kota/SN numerik yang ter-render float
        s = s[:-2]
    return s or None


def read_sheet(wb, name):
    """Baris data sebagai dict. Membuang baris kosong DAN baris header yang diulang di
    tengah sheet — sheet 2026 KSO Tes mengulang headernya 10x (satu per grup Type Alat);
    kalau ikut terbaca, dia jadi 10 'alat' ber-SN "SN".

    Pencocokan header sengaja case-insensitive dan berbasis ambang, bukan kesamaan persis:
    header baris-1 menulis 'JUNI' sementara header ulangan di baris 73 menulis 'Juni'.
    Satu huruf beda itu cukup untuk meloloskan seluruh baris kalau dibandingkan apa adanya.
    """
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    header_set = {h.upper() for h in header if h}
    out = []
    for row in it:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        nilai = [str(v).strip() for v in row if v is not None and str(v).strip()]
        cocok = sum(1 for v in nilai if v.upper() in header_set)
        if nilai and cocok >= max(3, len(nilai) * 0.6):
            continue
        out.append(dict(zip(header, row)))
    return out


def build(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    for s in (SHEET_POPULASI, SHEET_TES_2026, SHEET_REAGENT_2026):
        if s not in wb.sheetnames:
            sys.exit(f"Sheet wajib '{s}' tidak ada di workbook. Sheet tersedia: {wb.sheetnames}")

    report = {"sumber": os.path.basename(path), "sheet": {}, "peringatan": []}
    assets: dict[str, dict] = {}

    def layak(row, customer_field, alat_field):
        """Baris tanpa nama customer DAN tanpa nama alat tidak bisa jadi aset apa pun —
        itu sisa baris kosong yang cuma menyimpan satu sel yatim."""
        return bool(text(row.get(customer_field)) or text(row.get(alat_field)))

    def kunci(row, customer_field, alat_field):
        sn = norm_sn(row.get("SN"))
        if sn:
            return sn, text(row.get("SN"))
        # Tanpa SN, alat tetap harus punya identitas stabil supaya import idempoten.
        return f"NOSN:{slug(row.get(customer_field))}|{slug(row.get(alat_field))}", None

    # --- 1. Populasi KSO = metadata kontrak (MOU, target, paket, ritme) ---------------
    pop = read_sheet(wb, SHEET_POPULASI)
    status_map = {"PER TEST": "PER_TEST", "PERTES": "PER_TEST", "BELI REAGEN": "BELI_REAGEN"}
    dibuang = 0
    for r in pop:
        if not layak(r, "Customer", "Nama Alat"):
            dibuang += 1
            continue
        k, sn_raw = kunci(r, "Customer", "Nama Alat")
        assets[k] = {
            "sn_key": k, "sn_raw": sn_raw,
            "customer_raw": text(r.get("Customer")) or text(r.get("Alamat Penagihan")) or "(tanpa nama)",
            "kota": text(r.get("Kota")), "station": None, "admin": None,
            "type_alat": text(r.get("Type Alat")), "nama_alat": text(r.get("Nama Alat")),
            "skema": status_map.get(str(r.get("STATUS") or "").strip().upper(), "UNKNOWN"),
            "pemilik_alat": None,  # tidak ada sumbernya — lihat komentar migrasi 097
            "nomor_mou": text(r.get("Nomor MOU")),
            "mou_berlaku_sampai": as_date(r.get("Masa Berlaku MOU")),
            "target_jumlah_tes": as_int(r.get("Target Jumlah Tes")),
            "ritme_kunjungan": text(r.get("Ritme Kunjungan")),
            "paket": text(r.get("Paket")),
            "status_sheet": text(r.get("STATUS")),
            "keterangan": text(r.get("Keterangan")),
            "tgl_sj": text(r.get("TGL SJ")), "alamat": text(r.get("ALAMAT")),
            "outlet": text(r.get("OUTLET")),
            "in_populasi": True, "sumber_sheet": [SHEET_POPULASI], "catatan": [],
        }
    report["sheet"][SHEET_POPULASI] = {
        "baris": len(pop), "aset": len(assets), "baris_dibuang": dibuang}

    # --- 2. Sheet 2026 = kebenaran operasional untuk skema + station/admin ------------
    baru_2026 = 0
    for sheet, skema in ((SHEET_TES_2026, "PER_TEST"), (SHEET_REAGENT_2026, "BELI_REAGEN")):
        rows = read_sheet(wb, sheet)
        dibuang = 0
        for r in rows:
            if not layak(r, "Customer Real", "Nama Alat"):
                dibuang += 1
                continue
            k, sn_raw = kunci(r, "Customer Real", "Nama Alat")
            a = assets.get(k)
            if a is None:
                baru_2026 += 1
                a = assets[k] = {
                    "sn_key": k, "sn_raw": sn_raw,
                    "customer_raw": text(r.get("Customer Real")) or "(tanpa nama)",
                    "kota": text(r.get("Kota")), "station": None, "admin": None,
                    "type_alat": text(r.get("Type Alat")), "nama_alat": text(r.get("Nama Alat")),
                    "skema": skema, "pemilik_alat": None, "nomor_mou": None,
                    "mou_berlaku_sampai": None, "target_jumlah_tes": None,
                    "ritme_kunjungan": None, "paket": None, "status_sheet": None,
                    "keterangan": None, "tgl_sj": None, "alamat": None, "outlet": None,
                    "in_populasi": False, "sumber_sheet": [], "catatan": [],
                }
                a["catatan"].append(f"Tidak ada di {SHEET_POPULASI} — perlu disisir admin.")
            else:
                # Sheet 2026 menang atas STATUS di Populasi, tapi konfliknya dicatat.
                if a["skema"] not in ("UNKNOWN", skema):
                    a["catatan"].append(
                        f"Skema bentrok: STATUS Populasi '{a['status_sheet']}' "
                        f"({a['skema']}) vs keanggotaan sheet {sheet} ({skema}); "
                        f"dipakai {skema}.")
                elif a["skema"] == "UNKNOWN":
                    a["catatan"].append(f"STATUS kosong di Populasi; skema diambil dari {sheet}.")
                a["skema"] = skema
            if sheet in a["sumber_sheet"]:
                a["catatan"].append(f"SN muncul lebih dari sekali di {sheet}.")
            else:
                a["sumber_sheet"].append(sheet)
            a["station"] = a["station"] or text(r.get("Station"))
            a["admin"] = a["admin"] or text(r.get("Admin"))
            a["kota"] = a["kota"] or text(r.get("Kota"))
        report["sheet"][sheet] = {"baris": len(rows), "baris_dibuang": dibuang}

    for a in assets.values():
        if {SHEET_TES_2026, SHEET_REAGENT_2026} <= set(a["sumber_sheet"]):
            a["catatan"].append(
                "SN terdaftar di sheet Tes DAN Reagent 2026 — skema ganda, butuh keputusan.")

    # Skema tidak bisa ditentukan: STATUS kosong di Populasi DAN SN tak pernah muncul di
    # sheet Tes/Reagent mana pun. Sebelumnya kasus ini lolos tanpa catatan apa pun — hanya
    # cabang "UNKNOWN lalu ketemu di sheet" (di atas) yang ditandai, sementara yang tetap
    # UNKNOWN diam saja.
    #
    # KENAPA ITU BERBAHAYA: `kso_asset_produktivitas_v` mem-JOIN `kategori_skema`, yang cuma
    # mengenal PER_TEST & BELI_REAGEN. Baris UNKNOWN karena itu HILANG dari view — bukan
    # tampil dengan angka kosong, tapi lenyap. Pada impor prod pertama (2026-08-18) itu 22
    # aset, 10 di antaranya mesin Hemodialisa di RSUD Soegiri, dan tidak ada satu pun sinyal
    # di output yang menunjukkan mereka hilang: `catatan_sync`-nya NULL, jadi tidak ikut
    # terhitung di `aset_dengan_catatan` maupun tercetak di pratinjau importer.
    #
    # Catatan ini TIDAK menebak skemanya. Menyimpulkan dari alat lain milik faskes yang sama
    # justru bertentangan dengan desainnya sendiri: kolom `revenue_tumpang_tindih` ada persis
    # karena satu faskes bisa memegang dua skema sekaligus. Perbaikannya di sheet — isi
    # STATUS-nya — dan tugas skrip ini cuma memastikan kasusnya tidak lewat tanpa terlihat.
    skema_tak_tentu = 0
    for a in assets.values():
        if a["skema"] == "UNKNOWN":
            skema_tak_tentu += 1
            # DUA SEBAB, DUA TINDAKAN BERBEDA. Versi pertama catatan ini selalu berbunyi
            # "STATUS kosong" — keliru untuk baris yang STATUS-nya justru TERISI tapi nilainya
            # bukan skema. Pada data prod 2026-08-18 ada dua: 'BACKUP' (K Lyte 5, RSUD Ketapang)
            # dan 'NOT READY' (Fresenius, RSUD Soegiri). Keduanya status OPERASIONAL alat yang
            # tertulis di kolom jenis kerja sama. Menyuruh admin "mengisi STATUS" untuk baris
            # yang sudah ada isinya membuat catatan ini diabaikan, bukan ditindaklanjuti.
            #
            # Bedanya bukan kosmetik: yang kosong perlu DIISI, yang tak dikenali perlu
            # DIBETULKAN (nilainya salah kolom) — dan nilai seperti 'BACKUP' juga memberi tahu
            # kenapa alatnya nol realisasi tes.
            #
            # Salah ketik yang MEMANG bentuk skema ditangani di status_map (mis. 'PERTES'),
            # bukan di sini. Kalau muncul varian baru yang jelas-jelas berarti PER TEST atau
            # BELI REAGEN, tambahkan ke status_map; jangan menebaknya di catatan.
            st = (a.get("status_sheet") or "").strip()
            if st:
                a["catatan"].append(
                    f"Skema tidak dapat ditentukan: STATUS di Populasi berisi {st!r}, yang bukan "
                    "jenis kerja sama (dikenali: 'PER TEST', 'PERTES', 'BELI REAGEN'), dan SN "
                    "tidak terdaftar di sheet Tes maupun Reagent. Betulkan nilai STATUS-nya — "
                    "status operasional alat tempatnya di kolom Keterangan. Aset ini TIDAK akan "
                    "muncul di kso_asset_produktivitas_v sampai itu dibereskan.")
            else:
                a["catatan"].append(
                    "Skema tidak dapat ditentukan: STATUS kosong di Populasi dan SN tidak "
                    "terdaftar di sheet Tes maupun Reagent. Aset ini TIDAK akan muncul di "
                    "kso_asset_produktivitas_v sampai STATUS-nya diisi.")
    if skema_tak_tentu:
        report["peringatan"].append(
            f"{skema_tak_tentu} aset tidak punya skema (STATUS kosong ATAU tidak dikenali di "
            f"{SHEET_POPULASI}, dan tidak terdaftar di sheet Tes/Reagent). Aset-aset itu akan TERSIMPAN di kso_asset "
            "tapi TIDAK muncul di kso_asset_produktivitas_v. Perbaikannya di sheet: isi kolom "
            "STATUS yang kosong, dan BETULKAN yang nilainya bukan jenis kerja sama (status "
            "operasional alat seperti BACKUP/NOT READY tempatnya di kolom Keterangan). "
            "Jangan menebak dari alat lain milik faskes yang sama — satu faskes bisa "
            "memegang dua skema sekaligus.")

    # Sebagian "alat baru" sebetulnya alat lama yang SN-nya ditulis beda di sheet 2026
    # (mis. 5360 vs 05360-B, atau SN diketik ulang salah). Ditandai, TIDAK digabung otomatis:
    # menggabungkan dua SN berbeda tanpa konfirmasi bisa menghapus satu unit fisik dari
    # pembukuan. Admin yang memutuskan mana yang benar.
    by_cust = defaultdict(list)
    for a in assets.values():
        if a["in_populasi"]:
            by_cust[(slug(a["customer_raw"]), slug(a["nama_alat"]))].append(a["sn_key"])
    dugaan_duplikat = 0
    for a in assets.values():
        if a["in_populasi"]:
            continue
        kandidat = by_cust.get((slug(a["customer_raw"]), slug(a["nama_alat"])))
        if kandidat:
            dugaan_duplikat += 1
            a["catatan"].append(
                "Customer + nama alat sama persis dengan aset yang sudah ada di Populasi "
                f"(SN {', '.join(kandidat[:3])}) — kemungkinan SN-nya salah ketik, bukan unit baru.")

    # --- 3. Realisasi tes bulanan ----------------------------------------------------
    # 2025 dulu, 2026 belakangan: kalau ada tumpang tindih periode, angka dari sheet yang
    # tahunnya eksplisit sesuai periode yang menang (2026 di-load terakhir).
    tests: dict[tuple[str, str], dict] = {}
    non_numerik = Counter()
    tanpa_aset = Counter()
    for sheet in (SHEET_TES_2025, SHEET_REAGENT_2025, SHEET_TES_2026, SHEET_REAGENT_2026):
        if sheet not in wb.sheetnames:
            continue
        rows = read_sheet(wb, sheet)
        kolom = MONTH_COLUMNS[sheet]
        dipakai = 0
        for r in rows:
            if not layak(r, "Customer Real", "Nama Alat"):
                continue
            k, _ = kunci(r, "Customer Real", "Nama Alat")
            if k not in assets:
                # Sheet 2025 memuat unit yang sudah tidak ada di master 2026 (alat ditarik).
                # Angkanya dibuang: tidak ada aset yang bisa dijadikan induk.
                tanpa_aset[sheet] += 1
                continue
            for nama_kolom, (tahun, bulan) in kolom.items():
                if nama_kolom not in r:
                    continue
                mentah = r[nama_kolom]
                nilai = as_number(mentah)
                if nilai is None:
                    if mentah is not None and str(mentah).strip():
                        non_numerik[str(mentah).strip()[:20]] += 1
                    continue
                tests[(k, f"{tahun}-{bulan:02d}-01")] = {
                    "sn_key": k, "periode": f"{tahun}-{bulan:02d}-01",
                    "jumlah_tes": nilai, "sumber_sheet": sheet,
                }
                dipakai += 1
        report["sheet"].setdefault(sheet, {})["baris"] = len(rows)
        report["sheet"][sheet]["sel_bulanan_terpakai"] = dipakai
    report["sel_non_numerik"] = dict(non_numerik.most_common(15))
    report["baris_bulanan_tanpa_aset"] = dict(tanpa_aset)

    # --- 4. Per-parameter kimia klinik (sheet 2026, format long) ----------------------
    # Sheet ini memakai blok: baris pertama tiap alat berisi identitas, baris bulan
    # berikutnya identitasnya KOSONG (sel di-merge di Sheets). Tanpa forward-fill, 11 dari
    # 12 bulan tiap alat akan kehilangan induknya.
    params = []
    param_tanpa_aset = Counter()
    bulan_tak_dikenal = Counter()
    if SHEET_PARAM_2026 in wb.sheetnames:
        rows = read_sheet(wb, SHEET_PARAM_2026)
        konteks = {}
        for r in rows:
            if text(r.get("SN")) or text(r.get("Customer Real")):
                konteks = {
                    "SN": r.get("SN"), "Customer Real": r.get("Customer Real"),
                    "Nama Alat": r.get("Nama Alat"),
                }
            if not konteks:
                continue
            k, _ = kunci(konteks, "Customer Real", "Nama Alat")
            bulan_raw = r.get("BULAN")
            bulan = None
            if isinstance(bulan_raw, (datetime, date)):
                bulan = bulan_raw.month
            elif bulan_raw:
                s = str(bulan_raw).strip().upper()
                bulan = BULAN_KE_NOMOR.get(s) or BULAN_KE_NOMOR.get(s.split()[0])
            if not bulan:
                if bulan_raw:
                    bulan_tak_dikenal[str(bulan_raw)[:20]] += 1
                continue
            if k not in assets:
                param_tanpa_aset[k] += 1
                continue
            periode = f"2026-{bulan:02d}-01"
            for kolom, nama in PARAM_COLUMNS.items():
                nilai = as_number(r.get(kolom))
                if nilai is None:
                    continue
                params.append({
                    "sn_key": k, "periode": periode, "parameter": nama,
                    "jumlah_tes": nilai, "sumber_sheet": SHEET_PARAM_2026,
                })
        report["sheet"][SHEET_PARAM_2026] = {
            "baris": len(rows), "baris_parameter": len(params),
            "bulan_tak_dikenal": dict(bulan_tak_dikenal),
            "aset_tak_ketemu": len(param_tanpa_aset),
        }
        report["peringatan"].append(
            "Sheet `Detail Kimia` (tes_tertagih vs tes_di_alat, kontrol/standar/blank) BELUM "
            "di-import: sheet itu tidak punya kolom SN, hanya nama customer bebas — "
            "pencocokannya butuh peta manual dulu.")

    # --- 5. Rapikan + audit ----------------------------------------------------------
    for a in assets.values():
        a["catatan_sync"] = " | ".join(dict.fromkeys(a.pop("catatan"))) or None

    per_skema = Counter(a["skema"] for a in assets.values())
    report["total"] = {
        "aset": len(assets),
        "aset_in_populasi": sum(1 for a in assets.values() if a["in_populasi"]),
        "aset_di_luar_populasi": sum(1 for a in assets.values() if not a["in_populasi"]),
        "aset_tanpa_sn": sum(1 for a in assets.values() if a["sn_key"].startswith("NOSN:")),
        "aset_dengan_catatan": sum(1 for a in assets.values() if a["catatan_sync"]),
        # Hilang dari kso_asset_produktivitas_v sampai STATUS-nya diisi di sheet.
        "aset_skema_tak_tentu": skema_tak_tentu,
        "dugaan_duplikat_sn": dugaan_duplikat,
        "per_skema": dict(per_skema),
        "baris_tes_bulanan": len(tests),
        "baris_parameter": len(params),
    }
    # Aset yang ada di Populasi tapi tak sekalipun muncul di sheet 2026 — wajar untuk alat
    # yang tidak dihitung per-tes (Hemodialisa dll), tapi angkanya harus terlihat.
    diam = [a for a in assets.values() if a["in_populasi"] and len(a["sumber_sheet"]) == 1]
    report["total"]["aset_tanpa_realisasi_2026"] = len(diam)
    report["aset_tanpa_realisasi_2026_per_type"] = dict(
        Counter(a["type_alat"] or "(kosong)" for a in diam).most_common(15))

    return {
        "assets": list(assets.values()),
        "tests": list(tests.values()),
        "params": params,
        "report": report,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", help="path file .xlsx hasil export spreadsheet KSO")
    ap.add_argument("--out", required=True, help="path JSON keluaran (WAJIB di luar repo)")
    args = ap.parse_args()

    out = os.path.abspath(os.path.expanduser(args.out))
    repo = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    if out.startswith(repo + os.sep):
        sys.exit(f"Tolak menulis ke dalam repo ({repo}).\n"
                 "Repo ini publik dan JSON-nya memuat nama faskes, SN alat, dan nomor MOU.\n"
                 "Pilih path di luar repo, mis. ~/kso-import.json")

    hasil = build(os.path.expanduser(args.xlsx))
    with open(out, "w", encoding="utf-8") as f:
        json.dump(hasil, f, ensure_ascii=False, indent=1)

    print(json.dumps(hasil["report"], ensure_ascii=False, indent=2))
    print(f"\nJSON ditulis ke {out}")


if __name__ == "__main__":
    main()
