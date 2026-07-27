#!/usr/bin/env python3
"""wa-corpus-to-form.py — ubah hasil export-wa-corpus.sh menjadi draf
Form Input PIC Divisi (xlsx), memakai template resmi sebagai kerangka.

Mengisi sheet 'Daftar Posisi', 'A. Tugas & Target', 'B. Bedah SOP', dan
'C. Koordinasi' dengan kandidat dari korpus WhatsApp, lalu menambah sheet
'Bukti (auto WA)' berisi dasar tiap baris (jumlah pesan, grup, contoh pesan).
'OKR Divisi' TIDAK diisi — itu wewenang HOD.

Semua isian adalah DRAF untuk diverifikasi PIC, bukan fakta jobdesk/SOP.
Sel kuning = wajib diisi manusia (Level, Rules, Target, Target Level).

Pemakaian:
  python3 scripts/ops/wa-corpus-to-form.py \\
      --export ~/wa-corpus/export-wa-corpus-20260727-101500 \\
      --template "/path/Form-Input-PIC-Divisi_WRG-OS.xlsx"

Butuh openpyxl:  pip3 install openpyxl   (atau pakai services/ai/.venv)
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from copy import copy
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
except ImportError:
    sys.exit("butuh openpyxl → pip3 install openpyxl (atau jalankan pakai services/ai/.venv)")

# ── kerangka template ────────────────────────────────────────────────────────
# cols = jumlah kolom tabel · divisi/pic = sel judul baris 2 · first = baris data pertama
SHEETS = {
    "Daftar Posisi":     dict(cols=5, divisi="B2", pic="E2", first=6),
    "A. Tugas & Target": dict(cols=7, divisi="B2", pic="E2", first=6),
    "B. Bedah SOP":      dict(cols=5, divisi="B2", pic="E2", first=6),
    "C. Koordinasi":     dict(cols=4, divisi="B2", pic="D2", first=6),
}
HINT = "A4"
YELLOW = "FFFFF3C4"          # sel yang wajib diisi manusia (sama dgn template)
NEED_HUMAN = object()        # penanda: kosongkan + warnai kuning

# ── pemetaan posisi → divisi (bisa dioverride via --divisi-map) ──────────────
# Urutan penting: pola paling spesifik lebih dulu. Cocok sebagai kata utuh.
DEFAULT_MAP: list[tuple[str, str]] = [
    (r"technical service|teknisi|aftersales|engineering|engineer|kalibrasi", "Aftersales"),
    (r"accounting|tax|pajak", "Accounting"),
    (r"finance|treasury|invoic|piutang|\bar\b|\bcn\b|collection|kasir", "Finance & Supply Chain"),
    (r"supply chain|purchasing|procurement|inventory|gudang|warehouse|logistik|shipping|ekspedisi|admin sc",
     "Finance & Supply Chain"),
    (r"business analyst|\bivd\b|medical", "Business IVD & Medical"),
    (r"general affairs|\bga\b|\bit\b|\bhr\b|hrd|sdm|business development|\bbd\b|legal", "BD & GA"),
    (r"sales|account manager|\bam\b|kepala cabang|admin penjualan|kirim tagih|marketing",
     "Sales Area West & East"),
]
UNMAPPED = "BELUM DIPETAKAN"


def baca(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def angka(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return default


def frekuensi(hari: int, hari_ada_pesan: int) -> str:
    """Rasio hari-aktif → nilai dropdown Frekuensi template."""
    r = hari / max(hari_ada_pesan, 1)
    if r >= 0.60:
        return "Harian"
    if r >= 0.20:
        return "Mingguan"
    if r >= 0.05:
        return "Bulanan"
    return "Kuartalan"


def petakan(posisi: str, aturan: list[tuple[str, str]]) -> str:
    p = (posisi or "").lower()
    if not p or p == "tidak dikenal":
        return UNMAPPED
    for pola, divisi in aturan:
        if re.search(pola, p):
            return divisi
    return UNMAPPED


def potong(t: str, n: int) -> str:
    t = re.sub(r"\s+", " ", (t or "")).strip()
    return t if len(t) <= n else t[: n - 1].rstrip() + "…"


# ── penulisan sheet ─────────────────────────────────────────────────────────
def tulis_baris(ws, spec: dict, rows: list[list], gaya_ref_row: int) -> None:
    """Tulis rows mulai baris spec['first'], meniru gaya baris contoh."""
    first, ncol = spec["first"], spec["cols"]
    kuning = PatternFill("solid", fgColor=YELLOW)
    for i, data in enumerate(rows):
        r = first + i
        for c in range(1, ncol + 1):
            src = ws.cell(row=gaya_ref_row, column=c)
            dst = ws.cell(row=r, column=c)
            dst._style = copy(src._style)
            val = data[c - 1] if c - 1 < len(data) else None
            if val is NEED_HUMAN:
                dst.value = None
                dst.fill = kuning
            else:
                dst.value = val
    # bersihkan sisa baris contoh kalau data lebih pendek dari yang sudah terisi
    return None


def perluas_validasi(ws, baris_terakhir: int) -> None:
    """Panjangkan range dropdown supaya menutup semua baris draf."""
    for dv in ws.data_validations.dataValidation:
        baru = []
        for rng in list(dv.sqref.ranges):
            cr = CellRange(str(rng))
            if cr.max_row < baris_terakhir:
                cr = CellRange(
                    min_col=cr.min_col, min_row=cr.min_row,
                    max_col=cr.max_col, max_row=baris_terakhir,
                )
            baru.append(cr)
        dv.sqref = MultiCellRange(baru)


def set_judul(ws, spec: dict, divisi: str, pic: str, catatan: str) -> None:
    ws[spec["divisi"]] = divisi
    ws[spec["pic"]] = pic or ""
    ws[HINT] = catatan


# ── pembangun isi tiap sheet ────────────────────────────────────────────────
def isi_daftar_posisi(posisi_list, roster_per_posisi, wa_per_posisi):
    rows, bukti = [], []
    for n, pos in enumerate(posisi_list, 1):
        rr = roster_per_posisi.get(pos, [])
        orang_roster = sum(1 for r in rr if str(r.get("aktif", "")).lower() in ("t", "true", "1"))
        w = wa_per_posisi.get(pos, {})
        catatan = (
            f"WA: {w.get('pesan', 0)} pesan · {w.get('orang', 0)} orang terdeteksi"
            f" · grup: {potong(w.get('grup', ''), 60)}" if w else "tidak terdeteksi di korpus WA"
        )
        rows.append([n, pos, orang_roster or (w.get("orang") or None), NEED_HUMAN, catatan])
        bukti.append(["Daftar Posisi", pos, "roster master_user + 04_participation",
                      f"{orang_roster} aktif di roster; {w.get('pesan', 0)} pesan WA", ""])
    return rows, bukti


def isi_tugas(posisi_list, topik_per_posisi, pj_per_posisi, hari_ada_pesan, maks):
    rows, bukti = [], []
    for pos in posisi_list:
        for t in topik_per_posisi.get(pos, [])[:maks]:
            uraian = f"{t['topik']} — contoh: “{potong(t['contoh_pesan'], 110)}”"
            rows.append([
                pos, NEED_HUMAN, uraian, NEED_HUMAN,
                frekuensi(angka(t["hari"]), hari_ada_pesan),
                pj_per_posisi.get(pos) or NEED_HUMAN, NEED_HUMAN,
            ])
            bukti.append([
                "A. Tugas & Target", f"{pos} · {t['topik']}", "09b_topik_posisi.csv",
                f"{t['pesan']} pesan, {t['hari']}/{hari_ada_pesan} hari, {t['grup']} grup"
                f" · biasanya {t['hari_dominan']} ~{angka(t['jam_dominan'])}:00",
                potong(t["contoh_pesan"], 200),
            ])
    return rows, bukti


def isi_sop(alur_rows, posisi_set):
    rows, bukti = [], []
    for a in alur_rows:
        langkah = [s.strip() for s in a["alur_posisi"].split("→") if s.strip()]
        if len(langkah) < 2 or not (posisi_set & set(langkah)):
            continue
        topik = a.get("topik_dominan") or "(tanpa topik)"
        nama_sop = f"{topik} (draf dari alur WA)"
        catatan = (
            f"pola berulang {a['frekuensi']}× · rata {a['rata_pesan']} pesan / "
            f"{a['rata_durasi_menit']} mnt · grup: {potong(a['grup'], 50)}"
        )
        for i, pos in enumerate(langkah, 1):
            label = "(pengirim belum ter-resolve)" if pos == "TIDAK DIKENAL" else pos
            rows.append([
                nama_sop if i == 1 else None,
                f"{i}. {label}",
                "Manual",          # terjadi di WhatsApp → definisi Manual di Panduan
                NEED_HUMAN,
                catatan if i == 1 else None,
            ])
        bukti.append(["B. Bedah SOP", nama_sop, "11b_pola_alur.csv",
                      f"{a['frekuensi']}× · {a['jumlah_grup']} grup · alur: {a['alur_posisi']}",
                      f"pertama {a['pertama']} · terakhir {a['terakhir']}"])
    return rows, bukti


def isi_koordinasi(edge_rows, topik_pasangan, posisi_set, min_bobot):
    rows, bukti = [], []
    for e in edge_rows:
        if e["dari_posisi"] not in posisi_set:
            continue
        if angka(e["bobot"]) < min_bobot:
            continue
        if "TIDAK DIKENAL" in (e["dari_posisi"], e["ke_posisi"]):
            continue
        tp = topik_pasangan.get((e["dari_posisi"], e["ke_posisi"]), [])
        yang = " · ".join(t["topik"] for t in tp[:3]) or NEED_HUMAN
        jam = angka(e.get("jam_dominan"), -1)
        pemicu = (
            f"biasanya {e.get('hari_dominan') or '?'}"
            + (f" ~{jam}:00" if jam >= 0 else "")
            + f" · {e['bobot']}× interaksi / {e.get('hari_terjadi', '?')} hari"
        )
        rows.append([e["dari_posisi"], e["ke_posisi"], yang, pemicu])
        bukti.append([
            "C. Koordinasi", f"{e['dari_posisi']} → {e['ke_posisi']}",
            "06_matriks_posisi.csv + 06c_topik_pasangan_posisi.csv",
            f"bobot {e['bobot']} · balik {e['bobot_arah_balik']} · resiprositas {e['resiprositas']}"
            f" · grup: {potong(e['grup'], 50)}",
            potong(tp[0]["contoh_pesan"], 200) if tp else "",
        ])
    return rows, bukti


def tulis_bukti(wb, bukti: list[list], periode: str) -> None:
    ws = wb.create_sheet("Bukti (auto WA)")
    ws["A1"] = f"DASAR DRAF — diturunkan dari korpus WhatsApp {periode}"
    ws["A1"].font = copy(wb["Daftar Posisi"]["A5"].font)
    ws.append([])
    ws.append(["Sheet", "Baris draf", "Sumber berkas", "Angka pendukung", "Contoh pesan"])
    for c in ws[3]:
        c._style = copy(wb["Daftar Posisi"]["A5"]._style)
    for b in bukti:
        ws.append(b)
    for col, w in zip("ABCDE", (20, 42, 42, 60, 70)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# ── program utama ───────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--export", required=True, type=Path, help="folder hasil export-wa-corpus.sh")
    ap.add_argument("--template", required=True, type=Path, help="Form-Input-PIC-Divisi_WRG-OS.xlsx")
    ap.add_argument("--out", type=Path, help="folder tujuan (default <export>/form-pic)")
    ap.add_argument("--divisi-map", type=Path, help="CSV kolom: pola,divisi (override pemetaan default)")
    ap.add_argument("--only-divisi", help="hanya bangun satu divisi")
    ap.add_argument("--max-tugas", type=int, default=6, help="maks baris tugas per posisi (default 6)")
    ap.add_argument("--min-bobot", type=int, default=3, help="ambang bobot edge koordinasi (default 3)")
    ap.add_argument("--no-consolidated", action="store_true", help="jangan buat workbook SEMUA")
    a = ap.parse_args()

    exp: Path = a.export
    if not (exp / "00_MANIFEST.csv").exists():
        return print(f"✗ {exp} bukan folder hasil export (00_MANIFEST.csv tak ada)") or 2
    out = a.out or exp / "form-pic"
    out.mkdir(parents=True, exist_ok=True)

    # ── data ──
    param = (baca(exp / "00_parameter.csv") or [{}])[0]
    hari_ada_pesan = angka(param.get("hari_ada_pesan"), 1)
    periode = f"{param.get('pesan_pertama', '?')} .. {param.get('pesan_terakhir', '?')}"

    roster = baca(exp / "02_roster.csv")
    partisipasi = baca(exp / "04_participation.csv")
    edges = baca(exp / "06_matriks_posisi.csv")
    topik_pos = baca(exp / "09b_topik_posisi.csv")
    topik_pas = baca(exp / "06c_topik_pasangan_posisi.csv")
    alur = baca(exp / "11b_pola_alur.csv")

    roster_per_posisi = defaultdict(list)
    pj_per_posisi: dict[str, str] = {}
    atasan_tally = defaultdict(lambda: defaultdict(int))
    for r in roster:
        pos = (r.get("posisi") or "").strip()
        if not pos:
            continue
        roster_per_posisi[pos].append(r)
        atasan = (r.get("atasan_raw") or r.get("hod_key") or "").strip()
        if atasan:
            atasan_tally[pos][atasan] += 1
    for pos, tally in atasan_tally.items():
        pj_per_posisi[pos] = max(tally.items(), key=lambda kv: kv[1])[0]

    wa_per_posisi: dict[str, dict] = {}
    for p in partisipasi:
        pos = (p.get("position_key") or "").strip()
        d = wa_per_posisi.setdefault(pos, {"pesan": 0, "orang": set(), "grup": set()})
        d["pesan"] += angka(p.get("pesan"))
        d["orang"].add(p.get("person_key") or "")
        d["grup"].add(p.get("group_name") or "")
    for d in wa_per_posisi.values():
        d["orang"] = len({x for x in d["orang"] if x})
        d["grup"] = " | ".join(sorted(x for x in d["grup"] if x))

    topik_per_posisi = defaultdict(list)
    for t in sorted(topik_pos, key=lambda x: -angka(x.get("pesan"))):
        topik_per_posisi[(t.get("position_key") or "").strip()].append(t)

    topik_pasangan = defaultdict(list)
    for t in sorted(topik_pas, key=lambda x: -angka(x.get("bobot"))):
        topik_pasangan[(t["dari_posisi"], t["ke_posisi"])].append(t)

    # ── pemetaan divisi ──
    aturan = DEFAULT_MAP
    if a.divisi_map:
        aturan = [(r["pola"], r["divisi"]) for r in baca(a.divisi_map) if r.get("pola")]
        print(f"  pemetaan divisi dari {a.divisi_map} ({len(aturan)} aturan)")

    semua_posisi = sorted({p for p in list(roster_per_posisi) + list(wa_per_posisi) if p})
    divisi_posisi = defaultdict(list)
    with (out / "_pemetaan_divisi.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["posisi", "divisi", "orang_roster", "pesan_wa"])
        for pos in semua_posisi:
            div = petakan(pos, aturan)
            divisi_posisi[div].append(pos)
            w.writerow([pos, div, len(roster_per_posisi.get(pos, [])),
                        (wa_per_posisi.get(pos) or {}).get("pesan", 0)])

    # PIC & HOD dari template
    tpl = openpyxl.load_workbook(a.template)
    pic_of, hod_of = {}, {}
    ws_pic = tpl["PIC & HOD"]
    for r in ws_pic.iter_rows(min_row=5):
        divisi, pic, hod = (r[1].value, r[2].value, r[3].value)
        if divisi:
            pic_of[str(divisi).strip()] = str(pic or "").strip()
            hod_of[str(divisi).strip()] = str(hod or "").strip()

    target = list(divisi_posisi)
    if a.only_divisi:
        target = [d for d in target if d.lower() == a.only_divisi.lower()]
        if not target:
            return print(f"✗ divisi '{a.only_divisi}' tak ada. Pilihan: {', '.join(divisi_posisi)}") or 2
    if not a.no_consolidated:
        target.append("SEMUA")

    dibuat = []
    for divisi in target:
        posisi_list = semua_posisi if divisi == "SEMUA" else sorted(divisi_posisi[divisi])
        if not posisi_list:
            continue
        posisi_set = set(posisi_list)
        wb = openpyxl.load_workbook(a.template)
        bukti: list[list] = []
        catatan_umum = (
            f"DRAF OTOMATIS dari korpus WhatsApp {periode} ({param.get('total_pesan', '?')} pesan). "
            "Baris di bawah adalah kandidat untuk DIVERIFIKASI & DIKOREKSI PIC — bukan jobdesk/SOP resmi. "
            "Sel kuning wajib diisi manusia. Dasar tiap baris ada di sheet 'Bukti (auto WA)'."
        )

        data = {
            "Daftar Posisi": isi_daftar_posisi(posisi_list, roster_per_posisi, wa_per_posisi),
            "A. Tugas & Target": isi_tugas(posisi_list, topik_per_posisi, pj_per_posisi,
                                           hari_ada_pesan, a.max_tugas),
            "B. Bedah SOP": isi_sop(alur, posisi_set),
            "C. Koordinasi": isi_koordinasi(edges, topik_pasangan, posisi_set, a.min_bobot),
        }
        for nama, spec in SHEETS.items():
            ws = wb[nama]
            rows, bk = data[nama]
            bukti += bk
            set_judul(ws, spec, divisi, pic_of.get(divisi, ""), catatan_umum)
            if not rows:
                # jangan tinggalkan baris contoh template — bisa disalahpahami sbg data
                rows = [["(tidak ada kandidat dari korpus WhatsApp — isi manual)"]]
            tulis_baris(ws, spec, rows, gaya_ref_row=spec["first"])
            perluas_validasi(ws, spec["first"] + len(rows) - 1)
        # OKR: hanya isi identitas; isian tetap wewenang HOD
        okr = wb["OKR Divisi"]
        okr["B2"], okr["D2"] = divisi, pic_of.get(divisi, "")
        okr["A4"] = (f"Diisi HOD {hod_of.get(divisi, '')} — TIDAK diturunkan dari WhatsApp. "
                     "Target di sheet A menjadi Key Results-nya.")
        pand = wb["Panduan"]
        pand["B33"] = "Catatan draf otomatis (WRG-OS)"
        pand["B33"].font = copy(pand["B29"].font)
        pand["B34"] = (
            f"Sheet Daftar Posisi, A, B, dan C sudah berisi DRAF yang diturunkan dari korpus "
            f"WhatsApp {periode}. Draf ini kandidat, bukan jobdesk/SOP resmi — verifikasi, koreksi, "
            "dan hapus yang tak relevan. Sel kuning wajib diisi. Dasar tiap baris (jumlah pesan, "
            "grup, contoh pesan) ada di sheet 'Bukti (auto WA)'. OKR Divisi tetap diisi HOD."
        )
        pand["B34"].alignment = copy(pand["B30"].alignment)
        tulis_bukti(wb, bukti, periode)

        aman = re.sub(r"[^A-Za-z0-9]+", "-", divisi).strip("-")
        f = out / f"Form-PIC_{aman}_draf-WA.xlsx"
        wb.save(f)
        dibuat.append((divisi, f, sum(len(v[0]) for v in data.values())))

    print(f"\n  template : {a.template.name}")
    print(f"  periode  : {periode} · {hari_ada_pesan} hari ada pesan")
    for divisi, f, n in dibuat:
        print(f"  ✓ {divisi:<26} {n:>4} baris draf  → {f.name}")
    print(f"\n  pemetaan posisi→divisi: {out / '_pemetaan_divisi.csv'}")
    print("  (koreksi berkas itu jadi kolom pola,divisi lalu jalankan ulang dengan --divisi-map)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
