#!/usr/bin/env python3
"""Baca "Pemilik alat.xlsx" (LAPORAN PEMBELIAN KSO DARI PENYEDIA) → JSON siap-impor.

ATURAN YANG DITETAPKAN USER 2026-08-19:
    alat yang ADA di file ini  -> modalnya PENYEDIA, bukan WRG
    selebihnya yang ada di DB  -> milik WRG

Skrip ini hanya mengekstrak & mencocokkan; yang menulis ke DB adalah
scripts/ops/kso-pemilik-alat-apply.mjs. Alasan pemisahan sama dengan
kso-sheet-to-json.py: repo ini PUBLIK, dan keluarannya memuat nama faskes + SN alat,
jadi JSON-nya wajib ditulis di luar working tree (skrip menolak kalau tidak).

PAKAI:
    python3 scripts/ops/kso-pemilik-alat-to-json.py ~/Downloads/'Pemilik alat.xlsx' \
        --out ~/kso-pemilik.json
    node scripts/ops/kso-pemilik-alat-apply.mjs --file ~/kso-pemilik.json          # pratinjau
    node scripts/ops/kso-pemilik-alat-apply.mjs --file ~/kso-pemilik.json --apply
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl belum terpasang. Jalankan: python3 -m pip install openpyxl")

# Normalisasi SN HARUS identik dengan kso-sheet-to-json.py / migrasi 097. Kalau menyimpang,
# pencocokannya meleset tanpa error dan seluruh aset penyedia akan tercap milik WRG.
def norm_sn(raw) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = s.upper().replace(" ", "")
    return s.lstrip("0") or s


def build(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["KSO"] if "KSO" in wb.sheetnames else wb.worksheets[0]

    # Header ada di baris 4 — tiga baris pertama judul laporan. Dicari, bukan
    # dipatok, supaya laporan periode berikutnya yang judulnya lebih/kurang sebaris
    # tidak diam-diam terbaca sebagai data.
    baris_header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        nilai = [str(v).strip().upper() for v in row if v is not None]
        if any(v.startswith("SN ALAT") for v in nilai):
            baris_header = i
            break
    if baris_header is None:
        sys.exit("Tidak menemukan baris header (kolom 'SN ALAT'). Format berubah?")

    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[baris_header]]
    kol_sn = next(h for h in hdr if h.upper().startswith("SN ALAT"))

    rows = []
    for r in ws.iter_rows(min_row=baris_header + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        d = dict(zip(hdr, r))
        sn = norm_sn(d.get(kol_sn))
        if not sn:
            continue
        rows.append({
            "snKey": sn,
            "snRaw": str(d.get(kol_sn)).strip(),
            "merk": (str(d["Merk"]).strip() if d.get("Merk") else None),
            "namaBarang": (str(d["Nama Barang"]).strip() if d.get("Nama Barang") else None),
            "vendor": (str(d["Nama Vendor"]).strip() if d.get("Nama Vendor") else None),
            "peruntukan": (str(d["PERUNTUKAN CUSTOMER"]).strip() if d.get("PERUNTUKAN CUSTOMER") else None),
            "keterangan": (str(d["Keterangan"]).strip() if d.get("Keterangan") else None),
        })

    judul = " ".join(str(ws.cell(r, 1).value or "").strip() for r in (1, 2)).strip()
    return {
        "sumber": os.path.basename(path),
        "judul": judul,
        "penyedia": rows,
        "catatan": (
            "Baris di sini = alat bermodal PENYEDIA. Aset lain di kso_asset dianggap "
            "milik WRG sesuai keputusan user 2026-08-19."
        ),
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--out", required=True, help="path JSON keluaran (WAJIB di luar repo)")
    args = ap.parse_args()

    out = os.path.abspath(os.path.expanduser(args.out))
    repo = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    if out.startswith(repo + os.sep):
        sys.exit(f"Tolak menulis ke dalam repo ({repo}).\n"
                 "Repo ini publik dan JSON-nya memuat nama faskes serta SN alat.")

    hasil = build(os.path.expanduser(args.xlsx))
    with open(out, "w", encoding="utf-8") as f:
        json.dump(hasil, f, ensure_ascii=False, indent=1)

    print(f"{hasil['judul']}")
    print(f"baris penyedia: {len(hasil['penyedia'])}")
    ket = {}
    for r in hasil["penyedia"]:
        ket[r["keterangan"] or "(kosong)"] = ket.get(r["keterangan"] or "(kosong)", 0) + 1
    for k, v in sorted(ket.items(), key=lambda x: -x[1]):
        print(f"   {v:>3}  {k}")
    print(f"\nJSON ditulis ke {out}")


if __name__ == "__main__":
    main()
