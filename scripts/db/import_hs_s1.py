#!/usr/bin/env python3
"""F1-SPT importer — HS-S-1 Sales Activity Review (xlsx) → tabel `deal`.

One-shot, idempoten (NOT EXISTS pada facility+brand+product+am_id). Baca tab IVD
(product_category=IVD) & 'Non IVD' (Medical); abaikan 'Copy of IVD' & 'Non IVD,'.
Normalisasi Status→8-stage; derive kategori/probabilitas/forecast dari stage;
Sales Name→am_id (master_user.panggilan, alias VIKI→VICKY); PIC HOD→pic_hod;
cabang dari AM; account_id fuzzy facility→accurate_customer (pg_trgm ≥0.7, di SQL).

Pakai: python3 import_hs_s1.py --file /tmp/hs-s-1.xlsx [--db wrg_os_dev] [--apply]
  default = DRY-RUN (txn + ROLLBACK, cuma laporan; TIDAK insert).
"""
import argparse, csv, subprocess, sys, re, tempfile, os
from collections import Counter, defaultdict
from openpyxl import load_workbook

# ── mapping ────────────────────────────────────────────────────────────────
# PROGRES → 7-stage (migrasi 076: 'First Contact' lebur ke Prospecting,
# 'Offering' ke Quotation, tahap 'Closing' baru — tak ada padanan di sheet lama).
STATUS_MAP = {
    "lose": "Closing-Lost", "gagal": "Closing-Lost",
    "deal": "Closing-Won", "mou": "Closing-Won",
    "follow up": "Prospecting", "minat/tertarik": "Prospecting", "minat": "Prospecting",
    "sph": "Quotation",
    "offering letter": "Quotation", "offering": "Quotation",
    "presentation": "Presentation", "presentasi": "Presentation",
    "negotiating": "Negotiation", "negosiasi": "Negotiation", "proses di manajemen": "Negotiation",
    # ditunda ditangani khusus (on_hold)
}
# stage → (prospect_category, probability, forecast_category)
# HARUS sinkron dgn STAGE_META apps/api/src/repo/deal.ts + migrasi 076.
STAGE_DERIVE = {
    "Prospecting":   ("Cold", 0.10, "D - Omit"),
    "Presentation":  ("Cold", 0.30, "C - Pipeline"),
    "Quotation":     ("Cold", 0.50, "C - Pipeline"),
    "Negotiation":   ("Warm", 0.70, "B - Best Case"),
    "Closing":       ("Hot",  0.90, "A - Commit"),
    "Closing-Won":   ("Hot",  1.00, "Won"),
    "Closing-Lost":  ("",     0.00, "Lost"),
}
SALES_ALIAS = {"VIKI": "VICKY"}
MONTHS = {m: i for i, m in enumerate(
    ["januari","februari","maret","april","mei","juni","juli","agustus","september","oktober","november","desember"], 1)}
MONTHS.update({m: i for i, m in enumerate(
    ["january","february","march","april","may","june","july","august","september","october","november","december"], 1)})
INSTANSI = {"rumah sakit": "RS", "klinik": "Klinik", "puskesmas": "Puskesmas",
            "dinas kesehatan": "Dinkes", "laboratorium": "Lab", "laboratorium klinik": "Lab"}

def s(v):  # cell → stripped str
    return "" if v is None else str(v).strip()

def to_stage(status):
    st = s(status).lower()
    if st in ("ditunda",): return "Prospecting", True   # on_hold
    return STATUS_MAP.get(st, "Prospecting"), False

def parse_amount(v):
    if v is None: return ""
    if isinstance(v, (int, float)): return str(int(v))
    digits = re.sub(r"[^\d]", "", str(v))
    return digits if digits else ""

def parse_qty(v):
    t = s(v)
    if not t: return "", "", ""
    m = re.match(r"^\s*([\d.,]+)\s*(.*)$", t)
    if m:
        num = re.sub(r"[^\d]", "", m.group(1))
        return t, (num if num else ""), m.group(2).strip()
    return t, "", ""

def load_am_map(db):
    out = subprocess.run(["psql", db, "-tAF|", "-c",
        "select upper(btrim(panggilan)), am_id, coalesce(cabang,'') from master_user where coalesce(panggilan,'')<>''"],
        capture_output=True, text=True).stdout
    m = {}
    for line in out.strip().splitlines():
        p = line.split("|")
        if len(p) >= 3: m[p[0]] = (p[1], p[2])
    return m

# ── main ───────────────────────────────────────────────────────────────────
ap = argparse.ArgumentParser()
ap.add_argument("--file", required=True)
ap.add_argument("--db", default="wrg_os_dev")
ap.add_argument("--apply", action="store_true", help="commit (default dry-run rollback)")
args = ap.parse_args()

am_map = load_am_map(args.db)
wb = load_workbook(args.file, data_only=True)
TABS = {"IVD": "IVD", "Non IVD": "Medical"}

COLS = ["customer_name","facility_name","brand","product","parameter","prospect_category",
        "instansi_type","city","province","product_category","pic_hod","am_id","cabang",
        "coop_model","qty_text","qty_num","qty_unit","estimate_amount","pagu_info",
        "purchase_month","purchase_year","stage","probability","forecast_category",
        "on_hold","loss_status","notes"]

rows_out = []
rep = {"per_stage": Counter(), "unmapped_sales": Counter(), "cat_shift": 0,
       "minat": 0, "kosong": 0, "ditunda": 0, "total": 0, "per_tab": Counter()}

for tab, prodcat in TABS.items():
    ws = wb[tab]
    allrows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(allrows) if r and any(s(c) == "Brand" for c in r))
    hdr = [s(c) for c in allrows[hi]]
    def col(name, alt=None):
        for n in ([name] + ([alt] if alt else [])):
            if n in hdr: return hdr.index(n)
        return None
    ix = {k: col(*v) if isinstance(v, tuple) else col(v) for k, v in {
        "brand": "Brand", "product": "Product", "parameter": ("Parameter", "Parameter / mentioned"),
        "pcat": "Prospect Category", "fac": "Medical Facility Name", "inst": "Instansi",
        "city": "City", "prov": "Provience", "pichod": "PIC HOD", "sales": "Sales Name",
        "qty": "Qty", "coop": "Project Model Cooperation (KSO / Sale )",
        "est": "Estimate Purchase Amount", "pagu": "Pagu Information", "pt": "Purchase Time",
        "th": "Tahun", "status": "Status", "note": "NOTE"}.items()}
    for r in allrows[hi+1:]:
        if not r: continue
        brand = s(r[ix["brand"]]) if ix["brand"] is not None and len(r) > ix["brand"] else ""
        fac = s(r[ix["fac"]]) if ix["fac"] is not None and len(r) > ix["fac"] else ""
        if not brand or not fac: continue   # baris valid = ada brand + facility
        rep["total"] += 1; rep["per_tab"][tab] += 1
        get = lambda k: s(r[ix[k]]) if ix[k] is not None and len(r) > ix[k] else ""
        status_raw = get("status")
        stage, on_hold = to_stage(status_raw)
        if status_raw.lower() in ("minat/tertarik", "minat"): rep["minat"] += 1
        if not status_raw: rep["kosong"] += 1
        if on_hold: rep["ditunda"] += 1
        rep["per_stage"][stage] += 1
        pcat_derived, prob, forecast = STAGE_DERIVE[stage]
        pcat_sheet = get("pcat")
        if pcat_sheet and pcat_derived and pcat_sheet.lower() != pcat_derived.lower(): rep["cat_shift"] += 1
        # sales → am_id
        sales = get("sales").upper()
        sales = SALES_ALIAS.get(sales, sales)
        am_id, cabang = "", ""
        if sales and sales != "VACANT":
            if sales in am_map: am_id, cabang = am_map[sales]
            else: rep["unmapped_sales"][sales] += 1
        elif sales == "VACANT": rep["unmapped_sales"]["VACANT"] += 1
        qty_text, qty_num, qty_unit = parse_qty(get("qty"))
        rows_out.append({
            "customer_name": fac, "facility_name": fac, "brand": brand, "product": get("product"),
            "parameter": get("parameter"), "prospect_category": pcat_derived,
            "instansi_type": INSTANSI.get(get("inst").lower(), get("inst")),
            "city": get("city"), "province": get("prov"), "product_category": prodcat,
            "pic_hod": get("pichod"), "am_id": am_id, "cabang": cabang,
            "coop_model": get("coop"), "qty_text": qty_text, "qty_num": qty_num, "qty_unit": qty_unit,
            "estimate_amount": parse_amount(r[ix["est"]] if ix["est"] is not None and len(r) > ix["est"] else None),
            "pagu_info": get("pagu"),
            "purchase_month": str(MONTHS.get(get("pt").lower(), "")) if get("pt") else "",
            "purchase_year": re.sub(r"[^\d]", "", get("th"))[:4],
            "stage": stage, "probability": str(prob), "forecast_category": forecast,
            "on_hold": "true" if on_hold else "false",
            "loss_status": "approved" if stage == "Closing-Lost" else "",
            "notes": get("note"),
        })

# tulis staging CSV
csv_path = tempfile.mktemp(suffix="_hs_s1.csv")
with open(csv_path, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=COLS); w.writeheader()
    for row in rows_out: w.writerow(row)

# SQL: temp staging + \copy + INSERT (fuzzy account_id pg_trgm + dedup NOT EXISTS)
nz = lambda c: f"NULLIF(s.{c},'')"
sql = f"""
CREATE TEMP TABLE stg ({', '.join(c+' TEXT' for c in COLS)});
\\copy stg FROM '{csv_path}' WITH (FORMAT csv, HEADER true)
INSERT INTO deal (
  customer_name, facility_name, brand, product, parameter, prospect_category, instansi_type,
  city, province, product_category, pic_hod, am_id, cabang, coop_model, qty_text, qty_num,
  qty_unit, estimate_amount, pagu_info, purchase_month, purchase_year, stage, probability,
  forecast_category, on_hold, loss_status, notes, account_id)
SELECT {nz('customer_name')}, {nz('facility_name')}, {nz('brand')}, {nz('product')}, {nz('parameter')},
  {nz('prospect_category')}, {nz('instansi_type')}, {nz('city')}, {nz('province')}, {nz('product_category')},
  {nz('pic_hod')}, {nz('am_id')}, {nz('cabang')}, {nz('coop_model')}, {nz('qty_text')},
  {nz('qty_num')}::numeric, {nz('qty_unit')}, {nz('estimate_amount')}::numeric, {nz('pagu_info')},
  {nz('purchase_month')}::smallint, {nz('purchase_year')}::smallint, s.stage::deal_stage,
  {nz('probability')}::numeric, {nz('forecast_category')}, s.on_hold::boolean, {nz('loss_status')},
  {nz('notes')},
  (SELECT ac.id FROM accurate_customer ac
     WHERE similarity(ac.name, s.facility_name) >= 0.7
     ORDER BY similarity(ac.name, s.facility_name) DESC LIMIT 1)
FROM stg s
WHERE NOT EXISTS (
  SELECT 1 FROM deal d WHERE d.facility_name = s.facility_name
    AND coalesce(d.brand,'') = coalesce(s.brand,'')
    AND coalesce(d.product,'') = coalesce(NULLIF(s.product,''),'')
    AND coalesce(d.am_id,'') = coalesce(NULLIF(s.am_id,''),''));
\\echo '--- LAPORAN (dalam txn) ---'
SELECT 'deal_terisi=' || count(*) FROM deal;
SELECT 'account_id_matched=' || count(*) FILTER (WHERE account_id IS NOT NULL) || '/' || count(*) FROM deal;
SELECT 'stage: ' || stage || ' = ' || count(*) FROM deal GROUP BY stage ORDER BY 1;
"""

body = "BEGIN;\n" + sql + ("\nCOMMIT;\n" if args.apply else "\nROLLBACK;\n")
print(f"== HS-S-1 importer ({'APPLY' if args.apply else 'DRY-RUN'}) → db={args.db} ==")
print(f"  baris valid dibaca: {rep['total']} ({dict(rep['per_tab'])})")
print(f"  per-stage (mapping): {dict(rep['per_stage'])}")
print(f"  category shift (sheet vs derived): {rep['cat_shift']}")
print(f"  flag: MINAT/TERTARIK={rep['minat']} | Status kosong={rep['kosong']} | DITUNDA(on_hold)={rep['ditunda']}")
print(f"  sales tak ter-map: {dict(rep['unmapped_sales'])}")
print(f"  staging csv: {csv_path} ({len(rows_out)} baris)")
print("== DB (staging load + insert + report; " + ("COMMIT" if args.apply else "ROLLBACK") + ") ==")
res = subprocess.run(["psql", args.db, "-v", "ON_ERROR_STOP=1"], input=body, capture_output=True, text=True)
sys.stdout.write(res.stdout)
if res.returncode != 0:
    sys.stderr.write(res.stderr); sys.exit(1)
os.unlink(csv_path)
