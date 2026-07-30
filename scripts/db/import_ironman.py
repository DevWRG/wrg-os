#!/usr/bin/env python3
"""F1-SPT importer — PROGRES TEAM IRONMAN 2026 (xlsx) → tabel `deal`.

Baca sheet '2026' (data riil ~171 baris; sisanya baris kosong). Idempoten
(NOT EXISTS pada facility+brand+product+am_id) — dedup lintas HS-S-1 & IRONMAN,
jadi aman digabung satu pipeline & aman diulang.

Mapping (mengikuti import_hs_s1.py):
- PROGRES → 8-stage (STATUS_MAP; 'proses di manajemen'→Negotiation; 'ditunda'→on_hold).
- prospect_category / probability / forecast_category DIDERIVE dari stage
  (STATUS CUSTOMER sheet cuma dijadikan cross-check → laporan cat_shift).
- SALES → am_id (master_user.panggilan); VACANT → am_id kosong.
- product_category DIDERIVE dari brand → product_pricelist.lini (Medical/IVD),
  fallback kosong (di SQL).
- account_id fuzzy facility → accurate_customer (pg_trgm ≥0.7, di SQL).
- Sumber ditandai '[IRONMAN 2026]' di kolom notes untuk keterlacakan.

Pakai: python3 import_ironman.py --file <IRONMAN.xlsx> --db <wrg_os_dev|wrg_os_prod> [--apply]
  default = DRY-RUN (BEGIN + ROLLBACK; cuma laporan, TIDAK insert).
  --db WAJIB disebut (hindari salah-target DB).
"""
import argparse, csv, subprocess, sys, re, tempfile, os
from collections import Counter
from openpyxl import load_workbook

SHEET = "2026"

# PROGRES → 7-stage (identik dengan import_hs_s1.py; migrasi 076).
STATUS_MAP = {
    "lose": "Closing-Lost", "gagal": "Closing-Lost",
    "deal": "Closing-Won", "mou": "Closing-Won",
    "follow up": "Prospecting", "minat/tertarik": "Prospecting", "minat": "Prospecting",
    "sph": "Quotation",
    "offering letter": "Quotation", "offering": "Quotation",
    "presentation": "Presentation", "presentasi": "Presentation",
    "negotiating": "Negotiation", "negosiasi": "Negotiation", "proses di manajemen": "Negotiation",
    # 'ditunda' ditangani khusus (on_hold)
}
STAGE_DERIVE = {
    "Prospecting":   ("Cold", 0.10, "D - Omit"),
    "Presentation":  ("Cold", 0.30, "C - Pipeline"),
    "Quotation":     ("Cold", 0.50, "C - Pipeline"),
    "Negotiation":   ("Warm", 0.70, "B - Best Case"),
    "Closing":       ("Hot",  0.90, "A - Commit"),
    "Closing-Won":   ("Hot",  1.00, "Won"),
    "Closing-Lost":  ("",     0.00, "Lost"),
}
MONTHS = {m: i for i, m in enumerate(
    ["januari","februari","maret","april","mei","juni","juli","agustus","september","oktober","november","desember"], 1)}
INSTANSI = {"rumah sakit": "RS", "klinik": "Klinik", "puskesmas": "Puskesmas",
            "dinas kesehatan": "Dinkes", "laboratorium": "Lab", "laboratorium klinik": "Lab"}


def s(v):
    return "" if v is None else str(v).strip()


def to_stage(progres):
    st = s(progres).lower()
    if st == "ditunda":
        return "Prospecting", True   # on_hold
    return STATUS_MAP.get(st, "Prospecting"), False


def parse_amount(v):
    if v is None: return ""
    if isinstance(v, (int, float)): return str(int(v))
    digits = re.sub(r"[^\d]", "", str(v))
    return digits if digits else ""


def parse_qty(v):
    t = s(v)
    if not t: return "", "", ""
    m = re.match(r"^\s*([\d.,]+)\s*(.*)$", t)
    if m:
        num = re.sub(r"[^\d]", "", m.group(1))
        return t, (num if num else ""), (m.group(2).strip() or "pasien/bln")
    return t, "", ""


def load_am_map(db):
    out = subprocess.run(["psql", db, "-tAF|", "-c",
        "select upper(btrim(panggilan)), am_id, coalesce(cabang,'') from master_user where coalesce(panggilan,'')<>''"],
        capture_output=True, text=True).stdout
    m = {}
    for line in out.strip().splitlines():
        p = line.split("|")
        if len(p) >= 3: m[p[0]] = (p[1], p[2])
    return m


ap = argparse.ArgumentParser()
ap.add_argument("--file", required=True)
ap.add_argument("--db", required=True, help="wrg_os_dev / wrg_os_prod (WAJIB disebut)")
ap.add_argument("--apply", action="store_true", help="commit (default dry-run rollback)")
args = ap.parse_args()

am_map = load_am_map(args.db)
wb = load_workbook(args.file, data_only=True)
if SHEET not in wb.sheetnames:
    sys.exit(f"sheet '{SHEET}' tak ditemukan. Sheet ada: {wb.sheetnames}")
ws = wb[SHEET]
allrows = list(ws.iter_rows(values_only=True))
try:
    hi = next(i for i, r in enumerate(allrows) if r and any(s(c).upper() == "NAMA INSTANSI" for c in r))
except StopIteration:
    sys.exit("header 'NAMA INSTANSI' tak ketemu di sheet 2026.")
hdr = [s(c) for c in allrows[hi]]


def col(*names):
    for n in names:                                   # exact match dulu
        for j, h in enumerate(hdr):
            if h.upper() == n.upper(): return j
    for n in names:                                   # prefix match fallback
        for j, h in enumerate(hdr):
            if h.upper().startswith(n.upper()): return j
    return None


ix = {
    "fac": col("NAMA INSTANSI"),
    "statuscust": col("STATUS CUSTOMER"),
    "inst": col("JENIS INSTANSI"),
    "city": col("KOTA/KAB", "KOTA"),
    "parameter": col("PEMERIKSAAN"),
    "product": col("Produk", "PRODUK"),
    "brand": col("BRAND"),
    "sales": col("SALES"),
    "coop": col("Model Kerja sama", "Model Kerjasama"),
    "progres": col("PROGRES", "PROGRESS"),
    "detail": col("DETAIL MODEL"),
    "qty": col("PASIEN / BLN", "PASIEN/BLN", "PASIEN"),
    "harga": col("INFO HARGA"),
    "est": col("ESTIMASI SALES/BLN", "ESTIMASI SALES"),
    "deadline": col("DEADLINE"),
    "pic": col("pic"),
    "note": col("NOTE"),
}

COLS = ["customer_name","facility_name","brand","product","parameter","prospect_category",
        "instansi_type","city","province","product_category","pic_hod","am_id","cabang",
        "coop_model","qty_text","qty_num","qty_unit","estimate_amount","pagu_info",
        "purchase_month","purchase_year","stage","probability","forecast_category",
        "on_hold","loss_status","notes"]

rows_out = []
rep = {"per_stage": Counter(), "unmapped_sales": Counter(), "cat_shift": 0,
       "ditunda": 0, "kosong": 0, "total": 0}


def get(r, k):
    j = ix.get(k)
    return s(r[j]) if j is not None and len(r) > j else ""


for r in allrows[hi + 1:]:
    if not r: continue
    fac = get(r, "fac")
    if not fac: continue                              # baris valid = ada NAMA INSTANSI
    rep["total"] += 1
    progres = get(r, "progres")
    stage, on_hold = to_stage(progres)
    if not progres: rep["kosong"] += 1
    if on_hold: rep["ditunda"] += 1
    rep["per_stage"][stage] += 1
    pcat_derived, prob, forecast = STAGE_DERIVE[stage]
    sc = get(r, "statuscust")
    if sc and pcat_derived and sc.lower() != pcat_derived.lower(): rep["cat_shift"] += 1
    sales = get(r, "sales").upper()
    am_id, cabang = "", ""
    if sales and sales != "VACANT":
        if sales in am_map: am_id, cabang = am_map[sales]
        else: rep["unmapped_sales"][sales] += 1
    elif sales == "VACANT":
        rep["unmapped_sales"]["VACANT"] += 1
    qty_text, qty_num, qty_unit = parse_qty(get(r, "qty"))
    detail, note = get(r, "detail"), get(r, "note")
    notes = ("[IRONMAN 2026] " + " · ".join(
        x for x in [note, ("Detail: " + detail) if detail else ""] if x)).strip()
    rows_out.append({
        "customer_name": fac, "facility_name": fac, "brand": get(r, "brand"), "product": get(r, "product"),
        "parameter": get(r, "parameter"), "prospect_category": pcat_derived,
        "instansi_type": INSTANSI.get(get(r, "inst").lower(), get(r, "inst")),
        "city": get(r, "city"), "province": "", "product_category": "",   # diderive di SQL dari brand
        "pic_hod": get(r, "pic"), "am_id": am_id, "cabang": cabang,
        "coop_model": get(r, "coop"), "qty_text": qty_text, "qty_num": qty_num, "qty_unit": qty_unit,
        "estimate_amount": parse_amount(r[ix["est"]] if ix["est"] is not None and len(r) > ix["est"] else None),
        "pagu_info": get(r, "harga"),
        "purchase_month": str(MONTHS.get(get(r, "deadline").lower(), "")) if get(r, "deadline") else "",
        "purchase_year": "2026",
        "stage": stage, "probability": str(prob), "forecast_category": forecast,
        "on_hold": "true" if on_hold else "false",
        "loss_status": "approved" if stage == "Closing-Lost" else "",
        "notes": notes,
    })

csv_path = tempfile.mktemp(suffix="_ironman.csv")
with open(csv_path, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=COLS); w.writeheader()
    for row in rows_out: w.writerow(row)

nz = lambda c: f"NULLIF(s.{c},'')"
sql = f"""
CREATE TEMP TABLE stg ({', '.join(c + ' TEXT' for c in COLS)});
\\copy stg FROM '{csv_path}' WITH (FORMAT csv, HEADER true)
INSERT INTO deal (
  customer_name, facility_name, brand, product, parameter, prospect_category, instansi_type,
  city, province, product_category, pic_hod, am_id, cabang, coop_model, qty_text, qty_num,
  qty_unit, estimate_amount, pagu_info, purchase_month, purchase_year, stage, probability,
  forecast_category, on_hold, loss_status, notes, account_id)
SELECT {nz('customer_name')}, {nz('facility_name')}, {nz('brand')}, {nz('product')}, {nz('parameter')},
  {nz('prospect_category')}, {nz('instansi_type')}, {nz('city')}, {nz('province')},
  COALESCE((SELECT p.lini FROM product_pricelist p WHERE upper(p.brand)=upper(s.brand)
              GROUP BY p.lini ORDER BY count(*) DESC LIMIT 1), NULLIF(s.product_category,'')),
  {nz('pic_hod')}, {nz('am_id')}, {nz('cabang')}, {nz('coop_model')}, {nz('qty_text')},
  {nz('qty_num')}::numeric, {nz('qty_unit')}, {nz('estimate_amount')}::numeric, {nz('pagu_info')},
  {nz('purchase_month')}::smallint, {nz('purchase_year')}::smallint, s.stage::deal_stage,
  {nz('probability')}::numeric, {nz('forecast_category')}, s.on_hold::boolean, {nz('loss_status')},
  {nz('notes')},
  (SELECT ac.id FROM accurate_customer ac
     WHERE similarity(ac.name, s.facility_name) >= 0.7
     ORDER BY similarity(ac.name, s.facility_name) DESC LIMIT 1)
FROM stg s
WHERE NOT EXISTS (
  SELECT 1 FROM deal d WHERE d.facility_name = s.facility_name
    AND coalesce(d.brand,'') = coalesce(s.brand,'')
    AND coalesce(d.product,'') = coalesce(NULLIF(s.product,''),'')
    AND coalesce(d.am_id,'') = coalesce(NULLIF(s.am_id,''),''));
\\echo '--- LAPORAN (dalam txn) ---'
SELECT 'deal_total=' || count(*) FROM deal;
SELECT 'ironman_baris=' || count(*) FROM deal WHERE notes LIKE '[IRONMAN%';
SELECT 'product_category_terisi=' || count(*) FILTER (WHERE product_category IS NOT NULL) || '/' || count(*) FROM deal WHERE notes LIKE '[IRONMAN%';
SELECT 'account_id_matched=' || count(*) FILTER (WHERE account_id IS NOT NULL) || '/' || count(*) FROM deal WHERE notes LIKE '[IRONMAN%';
SELECT 'stage: ' || stage || ' = ' || count(*) FROM deal WHERE notes LIKE '[IRONMAN%' GROUP BY stage ORDER BY 1;
"""

body = "BEGIN;\n" + sql + ("\nCOMMIT;\n" if args.apply else "\nROLLBACK;\n")
print(f"== IRONMAN importer ({'APPLY' if args.apply else 'DRY-RUN'}) → db={args.db} · sheet={SHEET} ==")
print(f"  baris valid dibaca: {rep['total']}")
print(f"  per-stage (mapping): {dict(rep['per_stage'])}")
print(f"  category shift (STATUS CUSTOMER vs derived): {rep['cat_shift']}")
print(f"  flag: DITUNDA(on_hold)={rep['ditunda']} | PROGRES kosong={rep['kosong']}")
print(f"  sales tak ter-map: {dict(rep['unmapped_sales'])}")
print(f"  staging csv: {csv_path} ({len(rows_out)} baris)")
print("== DB (staging + insert + report; " + ("COMMIT" if args.apply else "ROLLBACK") + ") ==")
res = subprocess.run(["psql", args.db, "-v", "ON_ERROR_STOP=1"], input=body, capture_output=True, text=True)
sys.stdout.write(res.stdout)
if res.returncode != 0:
    sys.stderr.write(res.stderr); sys.exit(1)
os.unlink(csv_path)
